#!/usr/bin/env python3
"""
Daily Product Leadership Job Digest — personalised for Sujeet Singh
Sources: LinkedIn, Remotive, Built In, RemoteOK, TimesJobs,
         Naukri, iimjobs, Shine.com, Indeed India, Foundit India,
         The Product Folks, Instahyre, Weekday, Cutshort,
         Michael Page India, Stanton Chase, Heidrick & Struggles,
         ABC Consultants, Propella/Native, Longhouse, SutraHR,
         Korn Ferry India
"""

import requests
from bs4 import BeautifulSoup
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders as email_encoders
from datetime import datetime, date
import json
import os
import time
import sys
import io
from typing import Optional, List, Dict, Tuple

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("⚠  openpyxl not installed — Excel attachment will be skipped. Run: pip3 install openpyxl")

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────
# Secrets are read from the environment so nothing sensitive is ever
# committed. In GitHub Actions these come from repository Secrets; for a
# local run, export them in your shell first. See README.md.
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT   = 587

DEFAULT_SENDER = "sujeetkumar0809@gmail.com"


def load_config(env=None):
    """Load email credentials from the environment.

    Returns a dict with ``sender``, ``receiver`` and ``app_pass``.
    ``receiver`` defaults to ``sender`` when unset; ``app_pass`` is an
    empty string when no password is configured (so ``main`` can skip
    sending cleanly instead of crashing).
    """
    env = os.environ if env is None else env
    sender = env.get("SENDER_EMAIL", DEFAULT_SENDER)
    return {
        "sender":   sender,
        "receiver": env.get("RECEIVER_EMAIL", sender),
        "app_pass": env.get("GMAIL_APP_PASS", ""),
    }


_cfg           = load_config()
SENDER_EMAIL   = _cfg["sender"]
RECEIVER_EMAIL = _cfg["receiver"]
GMAIL_APP_PASS = _cfg["app_pass"]

# ─────────────────────────────────────────────
# PROFILE — Sujeet's relevancy signals
# ─────────────────────────────────────────────

# Titles the scraper will collect (used to filter scraped results)
LEADERSHIP_KEYWORDS = [
    # Executive / VP
    "vp of product", "vp product", "vice president of product",
    "vice president, product",
    # Director
    "director of product", "director, product", "product director",
    "senior director of product", "senior director, product",
    # Head / Chief
    "head of product", "head, product",
    "chief product officer", "cpo",
    # Group / Staff / Principal
    "group product manager", "gpm",
    "staff product manager",
    "principal product manager",
    # Senior / Lead IC
    "senior product manager",
    "lead product manager",
    "product lead",
    # Catch-all (covers "Product Manager - Growth", "Product Manager, Platform" etc.)
    "product manager",
]

# Search query strings sent to each platform
# Grouped so each platform can pick relevant tiers
EXEC_TITLES = [
    "VP of Product", "Director of Product", "Head of Product",
    "Chief Product Officer", "Senior Director of Product",
]
SENIOR_TITLES = [
    "Group Product Manager", "Principal Product Manager",
    "Staff Product Manager",
]
IC_TITLES = [
    "Senior Product Manager", "Lead Product Manager",
]
ALL_SEARCH_TITLES = EXEC_TITLES + SENIOR_TITLES + IC_TITLES

# Industries / domains that are a strong fit
STRONG_INDUSTRIES = [
    "health", "wellness", "medical", "clinical", "pharma", "medtech",
    "insurance", "insurtech", "fintech", "finance", "banking", "bfsi",
]
GOOD_INDUSTRIES = [
    "saas", "b2b", "enterprise", "platform", "analytics", "ai ",
    "tech", "digital", "cloud", "data", "automation",
]

# Title keywords that signal domain match
DOMAIN_TITLE_KEYWORDS = [
    "growth", "platform", "analytics", "gtm", "integration",
    "ai", "data", "revenue", "b2b", "consumer", "health",
]

# Industries that are clearly irrelevant → penalise heavily
BAD_INDUSTRIES = [
    "gaming", "video game", "game studio", "esport",
    "ecommerce", "e-commerce", "online shopping",
    "retail store", "brick and mortar",
    "logistics", "supply chain", "warehouse", "fulfillment",
    "semiconductor", "hardware device",
    "real estate", "proptech property",
    "food delivery", "restaurant", "hospitality", "hotel",
    "travel agency", "airline", "automotive", "manufacturing",
    "oil", "mining", "agriculture",
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

TODAY = date.today().isoformat()


# ─────────────────────────────────────────────
# HELPERS — HTTP
# ─────────────────────────────────────────────

def safe_get(url: str, timeout: int = 15, **kwargs) -> Optional[requests.Response]:
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout, **kwargs)
        r.raise_for_status()
        return r
    except Exception as e:
        print(f"  ⚠ GET failed {url[:65]}… — {e}")
        return None


def safe_post(url: str, timeout: int = 15, **kwargs) -> Optional[requests.Response]:
    try:
        kwargs.setdefault("headers", HEADERS)   # caller can override headers
        r = requests.post(url, timeout=timeout, **kwargs)
        r.raise_for_status()
        return r
    except Exception as e:
        print(f"  ⚠ POST failed {url[:65]}… — {e}")
        return None


# ─────────────────────────────────────────────
# HELPERS — PARSING & DEDUP
# ─────────────────────────────────────────────

def is_leadership(title: str) -> bool:
    t = title.lower()
    return any(kw in t for kw in LEADERSHIP_KEYWORDS)


def parse_jsonld_jobs(soup: BeautifulSoup, source: str, fallback_url: str) -> List[Dict]:
    jobs = []
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            raw = json.loads(script.string or "")
            if isinstance(raw, list):
                items = raw
            elif raw.get("@type") == "ItemList":
                items = [e.get("item", {}) for e in raw.get("itemListElement", [])]
            else:
                items = [raw]
            for item in items:
                if item.get("@type") != "JobPosting":
                    continue
                title = item.get("title", "")
                if not is_leadership(title):
                    continue
                org  = item.get("hiringOrganization", {})
                loc  = item.get("jobLocation", {})
                addr = loc.get("address", {}) if isinstance(loc, dict) else {}
                jobs.append({
                    "title":    title,
                    "company":  org.get("name", "") if isinstance(org, dict) else "",
                    "location": addr.get("addressLocality", "") if isinstance(addr, dict) else "",
                    "url":      item.get("url", fallback_url),
                    "source":   source,
                    "date":     (item.get("datePosted") or TODAY)[:10],
                })
        except (json.JSONDecodeError, AttributeError):
            continue
    return jobs


def deduplicate(jobs: List[Dict]) -> List[Dict]:
    seen, out = set(), []
    for j in jobs:
        key = (j["title"].lower().strip(), j["company"].lower().strip())
        if key not in seen:
            seen.add(key)
            out.append(j)
    return out


# ─────────────────────────────────────────────
# RELEVANCY SCORING  (1 – 10)
# ─────────────────────────────────────────────

def score_job(job: Dict) -> int:
    title    = job.get("title", "").lower()
    company  = job.get("company", "").lower()
    location = job.get("location", "").lower()
    combined = f"{title} {company} {location}"

    score = 5.0

    # ── Seniority fit ──────────────────────────────────────────
    SENIOR = ["vp ", "vice president", "director", "head of", "chief",
               "group product", "gpm", "principal"]
    JUNIOR = ["associate", "junior", "entry level", "intern", "coordinator"]
    if any(s in title for s in SENIOR):
        score += 1.0
    if any(s in title for s in JUNIOR):
        score -= 2.0

    # ── Industry fit ───────────────────────────────────────────
    if any(k in combined for k in STRONG_INDUSTRIES):
        score += 2.0
    elif any(k in combined for k in GOOD_INDUSTRIES):
        score += 1.0
    if any(k in combined for k in BAD_INDUSTRIES):
        score -= 3.0

    # ── Domain keyword match in title ─────────────────────────
    domain_hits = sum(1 for k in DOMAIN_TITLE_KEYWORDS if k in title)
    score += min(1.0, domain_hits * 0.4)

    # ── Location preference (India / Remote > US onsite) ──────
    INDIA_CITIES = ["india", "bengaluru", "bangalore", "mumbai", "delhi",
                    "hyderabad", "pune", "chennai", "kolkata", "gurugram",
                    "gurgaon", "noida"]
    if any(c in location for c in INDIA_CITIES):
        score += 1.0
    elif any(k in location for k in ["remote", "hybrid", "anywhere", "worldwide"]):
        score += 0.5
    elif any(k in location for k in ["new york", "san francisco", "seattle",
                                      "austin", "boston", "united states", "usa"]):
        score -= 0.5   # US-only onsite is geographically harder

    return max(1, min(10, round(score)))


# ─────────────────────────────────────────────
# COUNTRY FLAG DETECTION
# ─────────────────────────────────────────────

def detect_country_flag(location: str) -> str:
    loc = location.lower()
    INDIA   = ["india", "bengaluru", "bangalore", "mumbai", "delhi", "hyderabad",
               "pune", "chennai", "kolkata", "gurugram", "gurgaon", "noida"]
    USA     = ["united states", "usa", "u.s.", "new york", "san francisco",
               "seattle", "austin", "boston", "chicago", "los angeles", "sf bay",
               "silicon valley", "new jersey", "california", "texas", "washington dc"]
    UK      = ["united kingdom", "london", "manchester", "uk", "england",
               "birmingham", "edinburgh", "bristol"]
    CANADA  = ["canada", "toronto", "vancouver", "montreal", "calgary", "ottawa"]
    SING    = ["singapore"]
    AUS     = ["australia", "sydney", "melbourne", "brisbane", "perth"]
    GER     = ["germany", "berlin", "munich", "hamburg", "frankfurt"]
    NL      = ["netherlands", "amsterdam", "rotterdam"]
    FR      = ["france", "paris", "lyon"]
    UAE     = ["uae", "dubai", "abu dhabi"]
    REMOTE  = ["remote", "worldwide", "anywhere", "global", "distributed"]

    if any(k in loc for k in INDIA):   return "🇮🇳"
    if any(k in loc for k in USA):     return "🇺🇸"
    if any(k in loc for k in UK):      return "🇬🇧"
    if any(k in loc for k in CANADA):  return "🇨🇦"
    if any(k in loc for k in SING):    return "🇸🇬"
    if any(k in loc for k in AUS):     return "🇦🇺"
    if any(k in loc for k in GER):     return "🇩🇪"
    if any(k in loc for k in NL):      return "🇳🇱"
    if any(k in loc for k in FR):      return "🇫🇷"
    if any(k in loc for k in UAE):     return "🇦🇪"
    if any(k in loc for k in REMOTE):  return "🌍"
    if not location.strip():           return "🌐"
    return "🌐"


# ─────────────────────────────────────────────
# WORK-MODE DETECTION
# ─────────────────────────────────────────────

def detect_work_mode(title: str, location: str) -> Tuple[str, str]:
    """Returns (emoji, label)"""
    combined = f"{title} {location}".lower()
    if "remote" in combined:
        return ("🏠", "Remote")
    if "hybrid" in combined:
        return ("🔀", "Hybrid")
    if any(k in combined for k in ["on-site", "onsite", "in-office", "in office",
                                    "on site", "office only"]):
        return ("🏢", "On-site")
    return ("🔀", "Hybrid")   # safe default assumption


# ─────────────────────────────────────────────
# SOURCE 1 — Remotive
# ─────────────────────────────────────────────

def fetch_remotive() -> List[Dict]:
    print("Fetching Remotive …")
    jobs: List[Dict] = []
    seen_ids: set = set()
    # Try multiple API endpoints — category slug changed over time
    api_endpoints = [
        "https://remotive.com/api/remote-jobs?category=product&limit=100",
        "https://remotive.com/api/remote-jobs?category=product-management&limit=100",
        "https://remotive.com/api/remote-jobs?search=product+manager&limit=50",
        "https://remotive.com/api/remote-jobs?search=director+product&limit=30",
    ]
    for api_url in api_endpoints:
        r = safe_get(api_url, timeout=15)
        if not r:
            continue
        try:
            data = r.json()
        except Exception:
            continue
        for j in data.get("jobs", []):
            jid = str(j.get("id", ""))
            if jid and jid in seen_ids:
                continue
            seen_ids.add(jid)
            title = j.get("title", "")
            if not is_leadership(title):
                continue
            pub = j.get("publication_date", "") or TODAY
            jobs.append({
                "title":    title,
                "company":  j.get("company_name", ""),
                "location": j.get("candidate_required_location", "Remote"),
                "url":      j.get("url", ""),
                "source":   "Remotive",
                "date":     pub[:10],
            })
    print(f"  → {len(jobs)} roles")
    return jobs


# ─────────────────────────────────────────────
# SOURCE 2 — LinkedIn  (India + Remote focus)
# ─────────────────────────────────────────────

def fetch_linkedin() -> List[Dict]:
    print("Fetching LinkedIn …")
    # f_WT=2 = Remote  |  geoId=102713980 = India
    # Build query sets from shared title lists
    india_titles = EXEC_TITLES + SENIOR_TITLES + IC_TITLES
    remote_titles = EXEC_TITLES + SENIOR_TITLES + IC_TITLES
    query_sets = (
        [(t.replace(" ", "+"), "&geoId=102713980") for t in india_titles] +
        [(t.replace(" ", "+"), "&f_WT=2")          for t in remote_titles]
    )
    jobs = []
    for q, geo in query_sets:
        url = (
            f"https://www.linkedin.com/jobs/search/"
            f"?keywords={q}&f_TPR=r86400{geo}&position=1&pageNum=0"
        )
        r = safe_get(url, timeout=20)
        if not r:
            continue
        soup = BeautifulSoup(r.text, "html.parser")
        for card in soup.select("div.base-card"):
            title_el   = card.select_one("h3.base-search-card__title")
            company_el = card.select_one("h4.base-search-card__subtitle")
            loc_el     = card.select_one("span.job-search-card__location")
            link_el    = card.select_one("a.base-card__full-link")
            if not title_el:
                continue
            t = title_el.get_text(strip=True)
            if not is_leadership(t):
                continue
            href = (link_el.get("href", "") if link_el else "").split("?")[0]
            jobs.append({
                "title":    t,
                "company":  company_el.get_text(strip=True) if company_el else "",
                "location": loc_el.get_text(strip=True) if loc_el else "",
                "url":      href,
                "source":   "LinkedIn",
                "date":     TODAY,
            })
        time.sleep(1)
    print(f"  → {len(jobs)} roles")
    return jobs


# ─────────────────────────────────────────────
# SOURCE 3 — Built In
# ─────────────────────────────────────────────

def fetch_builtin() -> List[Dict]:
    print("Fetching Built In …")
    # Built In is a Next.js SPA — primary strategy: extract __NEXT_DATA__ JSON
    search_urls = [
        "https://builtin.com/jobs?title=Senior+Product+Manager&remote=true",
        "https://builtin.com/jobs?title=Director+of+Product&remote=true",
        "https://builtin.com/jobs?title=VP+Product&remote=true",
        "https://builtin.com/jobs?title=Head+of+Product&remote=true",
        "https://builtin.com/jobs?title=Principal+Product+Manager&remote=true",
    ]
    jobs: List[Dict] = []
    for url in search_urls:
        r = safe_get(url, timeout=20)
        if not r:
            time.sleep(0.5)
            continue
        soup = BeautifulSoup(r.text, "html.parser")
        # Strategy 1: JSON-LD
        jobs.extend(parse_jsonld_jobs(soup, "Built In", url))
        # Strategy 2: __NEXT_DATA__ embedded JSON (Next.js)
        script = soup.find("script", id="__NEXT_DATA__")
        if script:
            try:
                data = json.loads(script.string or "{}")
                pp = data.get("props", {}).get("pageProps", {})
                job_list = (
                    pp.get("jobs") or
                    pp.get("initialJobs") or
                    pp.get("jobListings") or
                    []
                )
                for j in job_list:
                    title = j.get("title") or j.get("name", "")
                    if not is_leadership(title):
                        continue
                    co = j.get("company") or {}
                    jurl = j.get("url", "") or j.get("slug", "")
                    if jurl and not jurl.startswith("http"):
                        jurl = "https://builtin.com" + jurl
                    jobs.append({
                        "title":    title,
                        "company":  co.get("name", "") if isinstance(co, dict) else str(co),
                        "location": j.get("locationStr") or j.get("location") or "Remote",
                        "url":      jurl or url,
                        "source":   "Built In",
                        "date":     (j.get("datePosted") or TODAY)[:10],
                    })
            except Exception:
                pass
        # Strategy 3: HTML data-cy attributes (Built In uses these)
        for a in soup.select("a[data-cy='job-title-link'], a.job-title, h2 a[href*='/job/']"):
            t = a.get_text(strip=True)
            if not t or not is_leadership(t):
                continue
            href = a.get("href", "")
            if not href.startswith("http"):
                href = "https://builtin.com" + href
            card = a.find_parent("li") or a.find_parent("div")
            c_el = card.select_one("[data-cy='job-company-name'], .company-name") if card else None
            l_el = card.select_one("[data-cy='job-location'], .job-location") if card else None
            jobs.append({
                "title":    t,
                "company":  c_el.get_text(strip=True) if c_el else "",
                "location": l_el.get_text(strip=True) if l_el else "Remote",
                "url":      href,
                "source":   "Built In",
                "date":     TODAY,
            })
        time.sleep(0.5)
    print(f"  → {len(jobs)} roles")
    return jobs


# ─────────────────────────────────────────────
# SOURCE 4 — RemoteOK  (replaces Otta which is now dead/redirected)
# Free public JSON API — no auth required
# ─────────────────────────────────────────────

def fetch_remoteok() -> List[Dict]:
    print("Fetching RemoteOK …")
    jobs: List[Dict] = []
    # tags=product-manager filters to PM roles; fallback to all and filter by title
    for api_url in [
        "https://remoteok.com/api?tags=product-manager",
        "https://remoteok.com/api?tags=product",
    ]:
        r = safe_get(api_url, timeout=20)
        if not r:
            continue
        try:
            items = r.json()
        except Exception:
            continue
        for j in items:
            # First item is a legal/metadata dict without "position" key — skip it
            if not isinstance(j, dict) or "position" not in j:
                continue
            title = j.get("position", "")
            if not is_leadership(title):
                continue
            raw_date = j.get("date", "") or TODAY
            jobs.append({
                "title":    title,
                "company":  j.get("company", ""),
                "location": j.get("location", "") or "Remote",
                "url":      j.get("url", "") or f"https://remoteok.com/remote-jobs/{j.get('id','')}",
                "source":   "RemoteOK",
                "date":     raw_date[:10],
            })
        if jobs:
            break   # first endpoint worked — don't double-count
    print(f"  → {len(jobs)} roles")
    return jobs


# ─────────────────────────────────────────────
# SOURCE 5 — TimesJobs  (replaces Google Jobs which blocks scrapers)
# Server-rendered HTML — reliable scraping target
# ─────────────────────────────────────────────

def fetch_timesjobs() -> List[Dict]:
    print("Fetching TimesJobs …")
    # Use top titles to avoid rate-limiting (server-rendered, safe to scrape)
    queries = [t.replace(" ", "+") for t in ALL_SEARCH_TITLES[:7]]
    jobs: List[Dict] = []
    for q in queries:
        url = (
            "https://www.timesjobs.com/candidate/job-search.html"
            f"?searchType=personalizedSearch&from=submit"
            f"&txtKeywords={q}&txtLocation=India"
        )
        r = safe_get(url, timeout=20)
        if not r:
            time.sleep(1)
            continue
        soup = BeautifulSoup(r.text, "html.parser")
        # JSON-LD (some listings include it)
        jobs.extend(parse_jsonld_jobs(soup, "TimesJobs", url))
        # TimesJobs job card selectors
        for card in soup.select("li.clearfix.job-bx, li[class*='job-bx'], ul.new-joblist li"):
            t_el = card.select_one("h2 a, .job-title a, header h2 a")
            c_el = card.select_one("h3.joblist-comp-name, .comp-name, h3 a")
            l_el = card.select_one("ul.top-jd-dtl li:first-child, .srp-skills span, .loc-name")
            if not t_el:
                continue
            t = t_el.get_text(strip=True)
            if not is_leadership(t):
                continue
            href = t_el.get("href", url)
            if not href.startswith("http"):
                href = "https://www.timesjobs.com" + href
            jobs.append({
                "title":    t,
                "company":  c_el.get_text(strip=True) if c_el else "",
                "location": l_el.get_text(strip=True) if l_el else "India",
                "url":      href,
                "source":   "TimesJobs",
                "date":     TODAY,
            })
        time.sleep(0.8)
    print(f"  → {len(jobs)} roles")
    return jobs


# ══════════════════════════════════════════════
# INDIA SOURCES
# ══════════════════════════════════════════════

# ─────────────────────────────────────────────
# SOURCE 6 — Naukri  (mass-market portal)
# ─────────────────────────────────────────────

def fetch_naukri() -> List[Dict]:
    print("Fetching Naukri …")
    # Use human-readable search keywords (not URL slugs) + map to SEO slug for HTML fallback
    keyword_map = [
        ("VP Product",               "vp-product-management-jobs"),
        ("Director of Product",      "director-product-management-jobs"),
        ("Head of Product",          "head-product-management-jobs"),
        ("Chief Product Officer",    "chief-product-officer-jobs"),
        ("Group Product Manager",    "group-product-manager-jobs"),
        ("Principal Product Manager","principal-product-manager-jobs"),
        ("Staff Product Manager",    "staff-product-manager-jobs"),
        ("Senior Product Manager",   "senior-product-manager-jobs"),
        ("Lead Product Manager",     "lead-product-manager-jobs"),
    ]
    jobs: List[Dict] = []
    api_headers = {
        **HEADERS,
        "appid":        "109",
        "systemid":     "Naukri",
        "content-type": "application/json",
        "Referer":      "https://www.naukri.com/",
    }
    for keyword, slug in keyword_map:
        k_enc = requests.utils.quote(keyword)
        api_url = (
            "https://www.naukri.com/jobapi/v3/search"
            f"?noOfResults=20&urlType=search_by_keyword&searchType=adv"
            f"&keyword={k_enc}&location=india&experience=5"
            f"&jobAge=30&src=jobsearchDesk&pageNo=1"
        )
        query_jobs: List[Dict] = []
        try:
            r = requests.get(api_url, headers=api_headers, timeout=15)
            if r.status_code == 200:
                data = r.json()
                job_list = data.get("jobDetails") or data.get("jobs") or []
                for j in job_list:
                    title = j.get("title", "")
                    if not is_leadership(title):
                        continue
                    placeholders = j.get("placeholders") or []
                    loc = ""
                    if placeholders and isinstance(placeholders[0], dict):
                        loc = placeholders[0].get("label", "")
                    query_jobs.append({
                        "title":    title,
                        "company":  j.get("companyName", ""),
                        "location": loc or "India",
                        "url":      j.get("jdURL", "") or f"https://www.naukri.com/{slug}",
                        "source":   "Naukri",
                        "date":     TODAY,
                    })
        except Exception as e:
            print(f"  ⚠ Naukri API [{keyword}]: {e}")

        # HTML fallback only if this keyword returned nothing
        if not query_jobs:
            r2 = safe_get(
                f"https://www.naukri.com/{slug}?l=india&experience=5",
                timeout=20
            )
            if r2:
                soup = BeautifulSoup(r2.text, "html.parser")
                query_jobs.extend(parse_jsonld_jobs(soup, "Naukri", f"https://www.naukri.com/{slug}"))
                for card in soup.select(
                    "article.jobTuple, div.jobTuple28, "
                    "div[class*='job-tuple'], div.srp-jobtuple-wrapper"
                ):
                    t_el = card.select_one("a.title, a[class*='title'], h2 a")
                    c_el = card.select_one(
                        "a.subTitle, .companyInfo a, [class*='comp-name'], .company-name"
                    )
                    l_el = card.select_one(".locWdth, [class*='location'], .loc")
                    t    = t_el.get_text(strip=True) if t_el else ""
                    if not t or not is_leadership(t):
                        continue
                    query_jobs.append({
                        "title":    t,
                        "company":  c_el.get_text(strip=True) if c_el else "",
                        "location": l_el.get_text(strip=True) if l_el else "India",
                        "url":      (t_el.get("href", "") if t_el else "") or f"https://www.naukri.com/{slug}",
                        "source":   "Naukri",
                        "date":     TODAY,
                    })
        jobs.extend(query_jobs)
        time.sleep(0.8)
    print(f"  → {len(jobs)} roles")
    return jobs


# ─────────────────────────────────────────────
# SOURCE 7 — iimjobs  (premium management portal)
# ─────────────────────────────────────────────

def fetch_iimjobs() -> List[Dict]:
    print("Fetching iimjobs …")
    # iimjobs is an Info Edge property (same group as Naukri)
    # Try their internal API with Naukri-style headers first
    api_headers = {
        **HEADERS,
        "appid":        "109",
        "systemid":     "iimjobs",
        "content-type": "application/json",
        "Referer":      "https://www.iimjobs.com/",
    }
    keywords = [
        "VP Product", "Director of Product", "Head of Product",
        "Senior Product Manager", "Group Product Manager", "Principal Product Manager",
    ]
    url_slugs = [
        "product-management-jobs?func=product-management&exp=10-30",
        "vice-president-product-jobs",
        "director-product-jobs",
        "senior-product-manager-jobs",
    ]
    jobs: List[Dict] = []

    # Strategy 1: Internal API (same as Naukri's)
    for kw in keywords:
        k_enc = requests.utils.quote(kw)
        api_url = (
            "https://www.iimjobs.com/jobapi/v3/search"
            f"?noOfResults=20&urlType=search_by_keyword&searchType=adv"
            f"&keyword={k_enc}&location=india&experience=5&pageNo=1"
        )
        try:
            r = requests.get(api_url, headers=api_headers, timeout=15)
            if r.status_code == 200:
                data = r.json()
                for j in (data.get("jobDetails") or data.get("jobs") or []):
                    title = j.get("title", "")
                    if not is_leadership(title): continue
                    placeholders = j.get("placeholders") or []
                    loc = placeholders[0].get("label", "India") if placeholders and isinstance(placeholders[0], dict) else "India"
                    jobs.append({
                        "title":    title,
                        "company":  j.get("companyName", ""),
                        "location": loc,
                        "url":      j.get("jdURL", f"https://www.iimjobs.com"),
                        "source":   "iimjobs",
                        "date":     TODAY,
                    })
        except Exception:
            pass
        time.sleep(0.4)

    # Strategy 2: HTML scraping their SEO pages (server-rendered)
    for slug in url_slugs:
        url = f"https://www.iimjobs.com/j/{slug}"
        r = safe_get(url, timeout=20)
        if not r:
            time.sleep(0.5)
            continue
        soup = BeautifulSoup(r.text, "html.parser")
        jobs.extend(parse_jsonld_jobs(soup, "iimjobs", url))
        # Try __NEXT_DATA__ (iimjobs may use Next.js)
        script = soup.find("script", id="__NEXT_DATA__")
        if script:
            try:
                data = json.loads(script.string or "{}")
                pp = data.get("props", {}).get("pageProps", {})
                for j in (pp.get("jobs") or pp.get("jobList") or []):
                    title = j.get("title") or j.get("designation", "")
                    if not is_leadership(title): continue
                    jobs.append({
                        "title":    title,
                        "company":  j.get("companyName", "") or j.get("company", ""),
                        "location": j.get("location", "") or j.get("city", "India"),
                        "url":      j.get("jobUrl", "") or j.get("url", url),
                        "source":   "iimjobs",
                        "date":     (j.get("postedOn", "") or TODAY)[:10],
                    })
            except Exception:
                pass
        # Multiple HTML selector attempts
        for card in soup.select(
            "li.job-wrapper, div.job-list-item, article[class*='job'], "
            "li[class*='job'], div[class*='job-card'], .jobItem, "
            "ul.job-list > li, div.job-details-wrapper"
        ):
            t_el = card.select_one(
                "h2 a, h3 a, a.job-title, a[class*='title'], "
                "a[href*='/job/'], .designation a, h1 a, .job-heading a"
            )
            c_el = card.select_one(
                ".company-name, a[class*='company'], .comp-name, "
                "[class*='companyName'], .organization, .emp-name"
            )
            l_el = card.select_one(".location, [class*='location'], .city, .job-location")
            if not t_el: continue
            t = t_el.get_text(strip=True)
            if not is_leadership(t): continue
            href = t_el.get("href", "")
            if not href.startswith("http"):
                href = "https://www.iimjobs.com" + href
            jobs.append({
                "title":    t,
                "company":  c_el.get_text(strip=True) if c_el else "",
                "location": l_el.get_text(strip=True) if l_el else "India",
                "url":      href,
                "source":   "iimjobs",
                "date":     TODAY,
            })
        time.sleep(0.5)
    print(f"  → {len(jobs)} roles")
    return jobs


# ─────────────────────────────────────────────
# SOURCE 8 — Shine.com  (replaces Hirist which is JS-rendered SPA)
# Shine is a major Indian portal with scrapeable pages + JSON-LD
# ─────────────────────────────────────────────

def fetch_shine() -> List[Dict]:
    print("Fetching Shine.com …")
    # Shine.com URL pattern for job search
    search_slugs = [
        "senior-product-manager-jobs",
        "director-of-product-jobs",
        "vp-product-jobs",
        "head-of-product-jobs",
        "principal-product-manager-jobs",
        "group-product-manager-jobs",
        "lead-product-manager-jobs",
    ]
    jobs: List[Dict] = []
    for slug in search_slugs:
        url = f"https://www.shine.com/job-search/{slug}/"
        r = safe_get(url, timeout=20)
        if not r:
            # Try alternate URL format
            q = slug.replace("-jobs", "").replace("-", "+")
            url = f"https://www.shine.com/job-search/?q={q}&l=India"
            r = safe_get(url, timeout=20)
        if not r:
            time.sleep(0.5)
            continue
        soup = BeautifulSoup(r.text, "html.parser")
        # JSON-LD structured data
        jobs.extend(parse_jsonld_jobs(soup, "Shine.com", url))
        # __NEXT_DATA__ (Shine uses Next.js)
        script = soup.find("script", id="__NEXT_DATA__")
        if script:
            try:
                data = json.loads(script.string or "{}")
                pp = data.get("props", {}).get("pageProps", {})
                job_list = (
                    pp.get("jobs") or pp.get("jobList") or
                    pp.get("searchResults", {}).get("jobs", []) or []
                )
                for j in job_list:
                    title = j.get("designation") or j.get("title", "")
                    if not is_leadership(title): continue
                    jobs.append({
                        "title":    title,
                        "company":  j.get("company", "") or j.get("companyName", ""),
                        "location": j.get("location", "") or j.get("city", "India"),
                        "url":      j.get("jobUrl", "") or j.get("url", url),
                        "source":   "Shine.com",
                        "date":     (j.get("postedOn", "") or TODAY)[:10],
                    })
            except Exception:
                pass
        # HTML fallback — Shine uses typical job card structure
        for card in soup.select(
            "div[class*='jobCard'], div[class*='job-card'], "
            "li[class*='job'], article[class*='job'], .job-listing-item"
        ):
            t_el = card.select_one(
                "h2 a, h3 a, a[class*='title'], a[class*='job-title'], "
                ".designation a, .job-heading a"
            )
            c_el = card.select_one("[class*='company'], [class*='org'], .comp-name")
            l_el = card.select_one("[class*='location'], [class*='city'], .loc")
            if not t_el: continue
            t = t_el.get_text(strip=True)
            if not is_leadership(t): continue
            href = t_el.get("href", "")
            if not href.startswith("http"):
                href = "https://www.shine.com" + href
            jobs.append({
                "title":    t,
                "company":  c_el.get_text(strip=True) if c_el else "",
                "location": l_el.get_text(strip=True) if l_el else "India",
                "url":      href or url,
                "source":   "Shine.com",
                "date":     TODAY,
            })
        time.sleep(0.5)
    print(f"  → {len(jobs)} roles")
    return jobs


# ─────────────────────────────────────────────
# SOURCE 9 — Indeed India  (aggregator)
# ─────────────────────────────────────────────

def fetch_indeed_india() -> List[Dict]:
    print("Fetching Indeed India …")
    # RSS feeds are far less blocked than HTML scraping on Indeed
    queries = [t.replace(" ", "+") for t in ALL_SEARCH_TITLES[:8]]
    jobs: List[Dict] = []
    for q in queries:
        # Primary: RSS feed
        rss_url = (
            f"https://in.indeed.com/rss?q={q}&l=India"
            f"&sort=date&fromage=14&limit=25"
        )
        r = safe_get(rss_url, timeout=20)
        if r and ("<rss" in r.text[:200] or "<item>" in r.text[:500]):
            soup = BeautifulSoup(r.content, "html.parser")
            for item in soup.find_all("item"):
                raw_title = (item.find("title") or {}).get_text(strip=True) if item.find("title") else ""
                # Indeed RSS titles: "Senior PM - Company Name"
                title = raw_title.split(" - ")[0].strip()
                if not is_leadership(title):
                    continue
                # Link in RSS is plain text after the <link> tag in Atom/RSS hybrid
                link_tag = item.find("link")
                href = ""
                if link_tag:
                    href = link_tag.get_text(strip=True)
                    if not href:
                        # sometimes it's next sibling
                        sib = link_tag.next_sibling
                        href = str(sib).strip() if sib else ""
                source_tag = item.find("source")
                company = source_tag.get_text(strip=True) if source_tag else ""
                jobs.append({
                    "title":    title,
                    "company":  company,
                    "location": "India",
                    "url":      href or rss_url,
                    "source":   "Indeed India",
                    "date":     TODAY,
                })
            time.sleep(1)
            continue

        # HTML fallback if RSS fails
        url = f"https://in.indeed.com/jobs?q={q}&l=India&sort=date&fromage=7&limit=20"
        r2 = safe_get(url, timeout=20)
        if not r2:
            time.sleep(1)
            continue
        soup = BeautifulSoup(r2.text, "html.parser")
        jobs.extend(parse_jsonld_jobs(soup, "Indeed India", url))
        for card in soup.select("div.job_seen_beacon, li[class*='css-']"):
            t_el = card.select_one("h2.jobTitle a, a[data-jk], h2 span")
            c_el = card.select_one("span[class*='companyName'], .companyName")
            l_el = card.select_one("div[class*='companyLocation'], .companyLocation")
            if not t_el:
                continue
            t = t_el.get_text(strip=True)
            if not is_leadership(t):
                continue
            href = t_el.get("href", "")
            if href and not href.startswith("http"):
                href = "https://in.indeed.com" + href
            jobs.append({
                "title":    t,
                "company":  c_el.get_text(strip=True) if c_el else "",
                "location": l_el.get_text(strip=True) if l_el else "India",
                "url":      href or url,
                "source":   "Indeed India",
                "date":     TODAY,
            })
        time.sleep(1)
    print(f"  → {len(jobs)} roles")
    return jobs


# ─────────────────────────────────────────────
# SOURCE 10 — Foundit India  (replaces Glassdoor India which requires JS + login)
# Foundit = Monster India rebranded. Try API + __NEXT_DATA__ + HTML fallback
# ─────────────────────────────────────────────

def fetch_foundit() -> List[Dict]:
    print("Fetching Foundit India …")
    jobs: List[Dict] = []
    search_terms = ALL_SEARCH_TITLES[:6]   # top 6 to avoid rate limiting

    for title in search_terms:
        q = requests.utils.quote(title)
        # Strategy 1: internal API endpoint
        api_url = (
            f"https://www.foundit.in/api/v1/jobs?query={q}"
            f"&location=India&experience=8&page=1&pageSize=20&sort=1"
        )
        r = safe_get(api_url, timeout=15)
        if r:
            try:
                data = r.json()
                job_list = (
                    data.get("jobsDto") or data.get("jobs") or
                    data.get("data", {}).get("jobs", []) or []
                )
                for j in job_list:
                    t = j.get("designation") or j.get("title", "")
                    if not is_leadership(t): continue
                    co = j.get("company") or {}
                    loc_list = j.get("locationList") or []
                    loc = loc_list[0].get("city", "India") if loc_list and isinstance(loc_list[0], dict) else j.get("location", "India")
                    jobs.append({
                        "title":    t,
                        "company":  co.get("name", "") if isinstance(co, dict) else str(co),
                        "location": loc,
                        "url":      j.get("jobDetailUrl") or j.get("url", api_url),
                        "source":   "Foundit India",
                        "date":     (j.get("postedOn", "") or TODAY)[:10],
                    })
                time.sleep(0.5)
                continue
            except Exception:
                pass

        # Strategy 2: HTML + __NEXT_DATA__
        page_url = (
            f"https://www.foundit.in/srp/results?query={q}"
            f"&location=India&experienceRanges=8~30&sort=1"
        )
        r2 = safe_get(page_url, timeout=20)
        if not r2:
            time.sleep(0.5)
            continue
        soup = BeautifulSoup(r2.text, "html.parser")
        jobs.extend(parse_jsonld_jobs(soup, "Foundit India", page_url))
        script = soup.find("script", id="__NEXT_DATA__")
        if script:
            try:
                data = json.loads(script.string or "{}")
                # Traverse common paths
                for path in [
                    ["props", "pageProps", "jobsDto"],
                    ["props", "pageProps", "jobs"],
                    ["props", "pageProps", "initialData", "jobs"],
                    ["props", "pageProps", "searchData", "jobs"],
                ]:
                    obj: object = data
                    for k in path:
                        obj = obj.get(k, {}) if isinstance(obj, dict) else {}
                    if isinstance(obj, list) and obj:
                        for j in obj:
                            t = j.get("designation") or j.get("title", "")
                            if not is_leadership(t): continue
                            jobs.append({
                                "title":    t,
                                "company":  j.get("companyName", "") or j.get("company", ""),
                                "location": j.get("location", "India"),
                                "url":      j.get("jobDetailUrl", page_url) or j.get("url", page_url),
                                "source":   "Foundit India",
                                "date":     TODAY,
                            })
                        break
            except Exception:
                pass
        time.sleep(0.8)
    print(f"  → {len(jobs)} roles")
    return jobs


# ─────────────────────────────────────────────
# SOURCE 11 — The Product Folks  (community)
# ─────────────────────────────────────────────

def fetch_tpf() -> List[Dict]:
    print("Fetching The Product Folks …")
    url = "https://www.theproductfolks.com/product-management-jobs"
    r = safe_get(url, timeout=20)
    if not r:
        return []
    soup = BeautifulSoup(r.text, "html.parser")
    jobs: List[Dict] = []
    jobs.extend(parse_jsonld_jobs(soup, "The Product Folks", url))
    for card in soup.select("div[class*='job'], article[class*='job'], li[class*='job']"):
        t_el = card.select_one("h2 a, h3 a, a[class*='title'], a[class*='job-title']")
        c_el = card.select_one("[class*='company'], [class*='org']")
        l_el = card.select_one("[class*='location'], [class*='city']")
        if not t_el:
            # fallback: any <a> whose text looks like a job title
            for a in card.select("a"):
                t = a.get_text(strip=True)
                if t and len(t) > 8 and is_leadership(t):
                    href = a.get("href", "")
                    if not href.startswith("http"):
                        href = "https://www.theproductfolks.com" + href
                    jobs.append({
                        "title": t, "company": "", "location": "India",
                        "url": href, "source": "The Product Folks", "date": TODAY,
                    })
            continue
        t = t_el.get_text(strip=True)
        if not is_leadership(t):
            continue
        href = t_el.get("href", "")
        if not href.startswith("http"):
            href = "https://www.theproductfolks.com" + href
        jobs.append({
            "title":    t,
            "company":  c_el.get_text(strip=True) if c_el else "",
            "location": l_el.get_text(strip=True) if l_el else "India",
            "url":      href,
            "source":   "The Product Folks",
            "date":     TODAY,
        })
    print(f"  → {len(jobs)} roles")
    return jobs


# ─────────────────────────────────────────────
# SOURCE 12 — Instahyre
# ─────────────────────────────────────────────

def fetch_instahyre() -> List[Dict]:
    print("Fetching Instahyre …")
    urls = [
        "https://www.instahyre.com/product-management-jobs/",
        "https://www.instahyre.com/search-jobs/?q=VP+Product&location=India",
        "https://www.instahyre.com/search-jobs/?q=Director+of+Product&location=India",
    ]
    jobs: List[Dict] = []
    for url in urls:
        r = safe_get(url, timeout=20)
        if not r:
            continue
        soup = BeautifulSoup(r.text, "html.parser")
        jobs.extend(parse_jsonld_jobs(soup, "Instahyre", url))
        for card in soup.select("div.opportunity-card, div[class*='job-card'], li[class*='job']"):
            t_el = card.select_one("h2 a, h3 a, a[class*='title'], a.job-title")
            c_el = card.select_one("[class*='company'], [class*='org-name']")
            l_el = card.select_one("[class*='location']")
            if not t_el:
                continue
            t = t_el.get_text(strip=True)
            if not is_leadership(t):
                continue
            href = t_el.get("href", "")
            if not href.startswith("http"):
                href = "https://www.instahyre.com" + href
            jobs.append({
                "title":    t,
                "company":  c_el.get_text(strip=True) if c_el else "",
                "location": l_el.get_text(strip=True) if l_el else "India",
                "url":      href,
                "source":   "Instahyre",
                "date":     TODAY,
            })
        time.sleep(0.5)
    print(f"  → {len(jobs)} roles")
    return jobs


# ─────────────────────────────────────────────
# SOURCE 13 — Weekday
# ─────────────────────────────────────────────

def fetch_weekday() -> List[Dict]:
    print("Fetching Weekday …")
    urls = [
        "https://weekday.works/jobs?roles=product-manager&location=India",
        "https://weekday.works/jobs?roles=vp-product&location=India",
    ]
    jobs: List[Dict] = []
    for url in urls:
        r = safe_get(url, timeout=20)
        if not r:
            continue
        soup = BeautifulSoup(r.text, "html.parser")
        jobs.extend(parse_jsonld_jobs(soup, "Weekday", url))
        for card in soup.select("div[class*='job'], li[class*='job'], article[class*='card']"):
            t_el = card.select_one("h2 a, h3 a, a[class*='title']")
            c_el = card.select_one("[class*='company']")
            l_el = card.select_one("[class*='location']")
            if not t_el:
                continue
            t = t_el.get_text(strip=True)
            if not is_leadership(t):
                continue
            href = t_el.get("href", "")
            if not href.startswith("http"):
                href = "https://weekday.works" + href
            jobs.append({
                "title":    t,
                "company":  c_el.get_text(strip=True) if c_el else "",
                "location": l_el.get_text(strip=True) if l_el else "India",
                "url":      href,
                "source":   "Weekday",
                "date":     TODAY,
            })
        time.sleep(0.5)
    print(f"  → {len(jobs)} roles")
    return jobs


# ─────────────────────────────────────────────
# SOURCE 14 — Cutshort
# ─────────────────────────────────────────────

def fetch_cutshort() -> List[Dict]:
    print("Fetching Cutshort …")
    urls = [
        "https://cutshort.io/jobs/product-management-jobs-in-india",
        "https://cutshort.io/jobs?q=VP+Product+Director+Product&location=India",
    ]
    jobs: List[Dict] = []
    for url in urls:
        r = safe_get(url, timeout=20)
        if not r:
            continue
        soup = BeautifulSoup(r.text, "html.parser")
        jobs.extend(parse_jsonld_jobs(soup, "Cutshort", url))
        for card in soup.select("div[class*='job-card'], li.job, article[class*='job']"):
            t_el = card.select_one("h2 a, h3 a, a[class*='title']")
            c_el = card.select_one("[class*='company'], [class*='startup']")
            l_el = card.select_one("[class*='location'], [class*='city']")
            if not t_el:
                continue
            t = t_el.get_text(strip=True)
            if not is_leadership(t):
                continue
            href = t_el.get("href", "")
            if not href.startswith("http"):
                href = "https://cutshort.io" + href
            jobs.append({
                "title":    t,
                "company":  c_el.get_text(strip=True) if c_el else "",
                "location": l_el.get_text(strip=True) if l_el else "India",
                "url":      href,
                "source":   "Cutshort",
                "date":     TODAY,
            })
        time.sleep(0.5)
    print(f"  → {len(jobs)} roles")
    return jobs


# ─────────────────────────────────────────────
# SOURCE 15 — Michael Page India  (executive search)
# ─────────────────────────────────────────────

def fetch_michael_page() -> List[Dict]:
    print("Fetching Michael Page India …")
    urls = [
        "https://www.michaelpage.co.in/jobs/technology/product",
        "https://www.michaelpage.co.in/jobs/technology/product-management",
        "https://www.michaelpage.co.in/jobs/human-resources-personnel/product-management",
    ]
    jobs: List[Dict] = []
    for url in urls:
        r = safe_get(url, timeout=20)
        if not r:
            continue
        soup = BeautifulSoup(r.text, "html.parser")
        jobs.extend(parse_jsonld_jobs(soup, "Michael Page India", url))
        for card in soup.select("article[class*='job'], div[class*='job-result'], li[class*='job']"):
            t_el = card.select_one("h2 a, h3 a, a[class*='job-title'], .job-title a")
            c_el = card.select_one("[class*='company'], [class*='client']")
            l_el = card.select_one("[class*='location'], [class*='city']")
            if not t_el:
                continue
            t = t_el.get_text(strip=True)
            if not is_leadership(t):
                continue
            href = t_el.get("href", "")
            if not href.startswith("http"):
                href = "https://www.michaelpage.co.in" + href
            jobs.append({
                "title":    t,
                "company":  c_el.get_text(strip=True) if c_el else "Confidential",
                "location": l_el.get_text(strip=True) if l_el else "India",
                "url":      href,
                "source":   "Michael Page India",
                "date":     TODAY,
            })
        time.sleep(0.5)
    print(f"  → {len(jobs)} roles")
    return jobs


# ─────────────────────────────────────────────
# SOURCE 16 — Stanton Chase India  (executive search)
# ─────────────────────────────────────────────

def fetch_stanton_chase() -> List[Dict]:
    print("Fetching Stanton Chase India …")
    url = "https://www.stantonchase.com/careers/?q=product+manager&location=India"
    r = safe_get(url, timeout=20)
    if not r:
        print("  → 0 roles")
        return []
    soup = BeautifulSoup(r.text, "html.parser")
    jobs = parse_jsonld_jobs(soup, "Stanton Chase", url)
    for card in soup.select("div[class*='career'], article[class*='job'], li[class*='position']"):
        t_el = card.select_one("h2 a, h3 a, a[class*='title']")
        l_el = card.select_one("[class*='location']")
        if not t_el:
            continue
        t = t_el.get_text(strip=True)
        if not is_leadership(t):
            continue
        href = t_el.get("href", "")
        if not href.startswith("http"):
            href = "https://www.stantonchase.com" + href
        jobs.append({
            "title": t, "company": "Confidential", "location": l_el.get_text(strip=True) if l_el else "India",
            "url": href, "source": "Stanton Chase", "date": TODAY,
        })
    print(f"  → {len(jobs)} roles")
    return jobs


# ─────────────────────────────────────────────
# SOURCE 17 — Heidrick & Struggles India  (executive search)
# ─────────────────────────────────────────────

def fetch_heidrick() -> List[Dict]:
    print("Fetching Heidrick & Struggles …")
    url = "https://www.heidrick.com/en/careers/job-search?q=product+manager&country=India"
    r = safe_get(url, timeout=20)
    if not r:
        print("  → 0 roles")
        return []
    soup = BeautifulSoup(r.text, "html.parser")
    jobs = parse_jsonld_jobs(soup, "Heidrick & Struggles", url)
    for card in soup.select("div[class*='job'], article[class*='position'], li[class*='vacancy']"):
        t_el = card.select_one("h2 a, h3 a, a[class*='title']")
        l_el = card.select_one("[class*='location']")
        if not t_el:
            continue
        t = t_el.get_text(strip=True)
        if not is_leadership(t):
            continue
        href = t_el.get("href", "")
        if not href.startswith("http"):
            href = "https://www.heidrick.com" + href
        jobs.append({
            "title": t, "company": "Confidential", "location": l_el.get_text(strip=True) if l_el else "India",
            "url": href, "source": "Heidrick & Struggles", "date": TODAY,
        })
    print(f"  → {len(jobs)} roles")
    return jobs


# ─────────────────────────────────────────────
# SOURCE 18 — ABC Consultants  (executive search)
# ─────────────────────────────────────────────

def fetch_abc_consultants() -> List[Dict]:
    print("Fetching ABC Consultants …")
    url = "https://www.abcconsultants.in/jobs/?s=product+manager"
    r = safe_get(url, timeout=20)
    if not r:
        print("  → 0 roles")
        return []
    soup = BeautifulSoup(r.text, "html.parser")
    jobs = parse_jsonld_jobs(soup, "ABC Consultants", url)
    for card in soup.select("article, div[class*='job'], li[class*='job']"):
        t_el = card.select_one("h2 a, h3 a, a[class*='title'], .entry-title a")
        l_el = card.select_one("[class*='location'], [class*='city']")
        if not t_el:
            continue
        t = t_el.get_text(strip=True)
        if not is_leadership(t):
            continue
        href = t_el.get("href", "")
        if not href.startswith("http"):
            href = "https://www.abcconsultants.in" + href
        jobs.append({
            "title": t, "company": "Confidential", "location": l_el.get_text(strip=True) if l_el else "India",
            "url": href, "source": "ABC Consultants", "date": TODAY,
        })
    print(f"  → {len(jobs)} roles")
    return jobs


# ─────────────────────────────────────────────
# SOURCE 19 — Propella / Native  (boutique tech search)
# ─────────────────────────────────────────────

def fetch_propella() -> List[Dict]:
    print("Fetching Propella/Native …")
    url = "https://propellasearch.com/leadership-hiring/"
    r = safe_get(url, timeout=20)
    if not r:
        print("  → 0 roles")
        return []
    soup = BeautifulSoup(r.text, "html.parser")
    jobs = parse_jsonld_jobs(soup, "Propella/Native", url)
    for card in soup.select("div[class*='job'], article[class*='position'], li[class*='role']"):
        t_el = card.select_one("h2 a, h3 a, a[class*='title']")
        l_el = card.select_one("[class*='location']")
        if not t_el:
            continue
        t = t_el.get_text(strip=True)
        if not is_leadership(t):
            continue
        href = t_el.get("href", "")
        if not href.startswith("http"):
            href = "https://propellasearch.com" + href
        jobs.append({
            "title": t, "company": "Confidential", "location": l_el.get_text(strip=True) if l_el else "India",
            "url": href, "source": "Propella/Native", "date": TODAY,
        })
    print(f"  → {len(jobs)} roles")
    return jobs


# ─────────────────────────────────────────────
# SOURCE 20 — Longhouse  (boutique)
# ─────────────────────────────────────────────

def fetch_longhouse() -> List[Dict]:
    print("Fetching Longhouse …")
    for url in ["https://longhouse.in/jobs/", "https://longhouse.in/careers/"]:
        r = safe_get(url, timeout=20)
        if not r:
            continue
        soup = BeautifulSoup(r.text, "html.parser")
        jobs = parse_jsonld_jobs(soup, "Longhouse", url)
        for a in soup.select("a[href*='job'], a[href*='career'], a[href*='role']"):
            t = a.get_text(strip=True)
            if t and len(t) > 8 and is_leadership(t):
                href = a.get("href", "")
                if not href.startswith("http"):
                    href = "https://longhouse.in" + href
                jobs.append({
                    "title": t, "company": "Confidential", "location": "India",
                    "url": href, "source": "Longhouse", "date": TODAY,
                })
        if jobs:
            print(f"  → {len(jobs)} roles")
            return jobs
    print("  → 0 roles")
    return []


# ─────────────────────────────────────────────
# SOURCE 21 — SutraHR  (boutique)
# ─────────────────────────────────────────────

def fetch_sutrahr() -> List[Dict]:
    print("Fetching SutraHR …")
    for url in ["https://www.sutrahr.com/jobs/", "https://www.sutrahr.com/current-openings/"]:
        r = safe_get(url, timeout=20)
        if not r:
            continue
        soup = BeautifulSoup(r.text, "html.parser")
        jobs = parse_jsonld_jobs(soup, "SutraHR", url)
        for card in soup.select("div[class*='job'], article, li[class*='opening']"):
            t_el = card.select_one("h2 a, h3 a, a[class*='title'], .entry-title a")
            if not t_el:
                continue
            t = t_el.get_text(strip=True)
            if not is_leadership(t):
                continue
            href = t_el.get("href", "")
            if not href.startswith("http"):
                href = "https://www.sutrahr.com" + href
            jobs.append({
                "title": t, "company": "Confidential", "location": "India",
                "url": href, "source": "SutraHR", "date": TODAY,
            })
        if jobs:
            print(f"  → {len(jobs)} roles")
            return jobs
    print("  → 0 roles")
    return []


# ─────────────────────────────────────────────
# SOURCE 22 — Korn Ferry India  (executive search)
# ─────────────────────────────────────────────

def fetch_korn_ferry_india() -> List[Dict]:
    print("Fetching Korn Ferry India …")
    for url in [
        "https://www.kornferry.com/jobs?search=product+manager&location=India",
        "https://www.kornferry.com/careers?q=product",
    ]:
        r = safe_get(url, timeout=20)
        if not r:
            continue
        soup = BeautifulSoup(r.text, "html.parser")
        jobs = parse_jsonld_jobs(soup, "Korn Ferry India", url)
        for card in soup.select("div[class*='job'], article[class*='position']"):
            t_el = card.select_one("h2 a, h3 a, a[class*='title']")
            l_el = card.select_one("[class*='location']")
            if not t_el:
                continue
            t = t_el.get_text(strip=True)
            if not is_leadership(t):
                continue
            href = t_el.get("href", "")
            if not href.startswith("http"):
                href = "https://www.kornferry.com" + href
            jobs.append({
                "title": t, "company": "Confidential", "location": l_el.get_text(strip=True) if l_el else "India",
                "url": href, "source": "Korn Ferry India", "date": TODAY,
            })
        if jobs:
            print(f"  → {len(jobs)} roles")
            return jobs
    print("  → 0 roles")
    return []


# ─────────────────────────────────────────────
# EMAIL BUILDER
# ─────────────────────────────────────────────

SOURCE_COLORS = {
    # Global sources
    "Remotive":            "#6366f1",
    "LinkedIn":            "#0077B5",
    "Built In":            "#0f766e",
    "RemoteOK":            "#e11d48",   # replaces Otta
    "TimesJobs":           "#ea4335",   # replaces Google Jobs
    # India mass-market
    "Naukri":              "#ff7555",
    "iimjobs":             "#2563eb",
    "Shine.com":           "#7c3aed",   # replaces Hirist
    "Indeed India":        "#003a9b",
    "Foundit India":       "#0caa41",   # replaces Glassdoor India
    # India community / tech
    "The Product Folks":   "#0891b2",
    "Instahyre":           "#d97706",
    "Weekday":             "#4f46e5",
    "Cutshort":            "#be185d",
    # Executive search
    "Michael Page India":  "#1e293b",
    "Stanton Chase":       "#374151",
    "Heidrick & Struggles":"#1f2937",
    "ABC Consultants":     "#312e81",
    "Propella/Native":     "#064e3b",
    "Longhouse":           "#78350f",
    "SutraHR":             "#7f1d1d",
    "Korn Ferry India":    "#1e293b",
}


def score_color(score: int) -> Tuple[str, str]:
    """Returns (background, text) hex colors for the score badge."""
    if score >= 8:
        return ("#dcfce7", "#166534")   # green
    if score >= 5:
        return ("#fef9c3", "#854d0e")   # amber
    return ("#fee2e2", "#991b1b")       # red


def source_badge(source: str) -> str:
    color = SOURCE_COLORS.get(source, "#64748b")
    return (
        f'<span style="background:{color};color:#fff;font-size:10px;'
        f'font-weight:700;padding:2px 7px;border-radius:20px;'
        f'white-space:nowrap">{source}</span>'
    )


def score_badge(score: int) -> str:
    bg, fg = score_color(score)
    return (
        f'<span style="background:{bg};color:{fg};font-size:12px;'
        f'font-weight:700;padding:3px 9px;border-radius:20px;'
        f'white-space:nowrap">⭐ {score}/10</span>'
    )


def build_html(jobs: List[Dict], sources_used: List[str]) -> str:
    today_str = datetime.now().strftime("%A, %B %d %Y")
    count     = len(jobs)

    by_source: Dict[str, List[Dict]] = {}
    for j in jobs:
        by_source.setdefault(j["source"], []).append(j)

    source_summary = " &nbsp;·&nbsp; ".join(
        f'<span style="color:{SOURCE_COLORS.get(s,"#64748b")};font-weight:700">'
        f'{s}&nbsp;({len(by_source.get(s,[]))})</span>'
        for s in sources_used if s in by_source
    )

    rows = ""
    for i, j in enumerate(jobs):
        bg           = "#ffffff" if i % 2 == 0 else "#f8fafc"
        score        = j.get("_score", 5)
        country_flag = j.get("_country_flag", "🌐")
        wm_emoji, wm_label = j.get("_work_mode", ("🔀", "Hybrid"))
        date_str     = j.get("date", "")

        rows += f"""
        <tr style="background:{bg}">
          <!-- Rank -->
          <td style="padding:12px 10px 12px 16px;border-bottom:1px solid #e2e8f0;
              color:#94a3b8;font-size:13px;font-weight:700;width:32px;
              text-align:center;vertical-align:top">{i+1}</td>

          <!-- Role + Company -->
          <td style="padding:12px 10px;border-bottom:1px solid #e2e8f0;vertical-align:top">
            <a href="{j['url']}" style="color:#1d4ed8;font-weight:700;font-size:14px;
               text-decoration:none;line-height:1.4">{j['title']}</a>
            <div style="margin-top:3px;color:#475569;font-size:12px">
              {'<strong style="color:#1e293b">' + j['company'] + '</strong>' if j['company'] else ''}
              {'&nbsp;&nbsp;' if j['company'] and date_str else ''}
              {'<span style="color:#94a3b8">' + date_str + '</span>' if date_str else ''}
            </div>
          </td>

          <!-- Country flag -->
          <td style="padding:12px 8px;border-bottom:1px solid #e2e8f0;
              text-align:center;vertical-align:top;white-space:nowrap">
            <span title="{j.get('location','')}" style="font-size:20px;
              cursor:default">{country_flag}</span>
          </td>

          <!-- Work mode -->
          <td style="padding:12px 8px;border-bottom:1px solid #e2e8f0;
              text-align:center;vertical-align:top;white-space:nowrap">
            <span title="{wm_label}" style="font-size:18px;cursor:default">{wm_emoji}</span>
            <div style="font-size:10px;color:#94a3b8;margin-top:1px">{wm_label}</div>
          </td>

          <!-- Relevancy score -->
          <td style="padding:12px 8px;border-bottom:1px solid #e2e8f0;
              text-align:center;vertical-align:top;white-space:nowrap">
            {score_badge(score)}
          </td>

          <!-- Source -->
          <td style="padding:12px 16px 12px 8px;border-bottom:1px solid #e2e8f0;
              vertical-align:top;white-space:nowrap">
            {source_badge(j['source'])}
          </td>
        </tr>"""

    no_jobs = '<tr><td colspan="6" style="padding:32px;text-align:center;color:#94a3b8">No new roles found today.</td></tr>'

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
</head>
<body style="margin:0;padding:0;background:#f1f5f9;
  font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Helvetica,Arial,sans-serif">

<table width="100%" cellpadding="0" cellspacing="0"
  style="background:#f1f5f9;padding:32px 16px">
<tr><td align="center">
<table width="780" cellpadding="0" cellspacing="0" style="max-width:780px;width:100%">

  <!-- Header -->
  <tr>
    <td style="background:linear-gradient(135deg,#1e40af 0%,#7c3aed 100%);
        border-radius:12px 12px 0 0;padding:24px 28px">
      <div style="color:#bfdbfe;font-size:11px;font-weight:700;letter-spacing:1.5px;
          text-transform:uppercase;margin-bottom:4px">Daily Digest · {today_str}</div>
      <h1 style="margin:0;color:#ffffff;font-size:22px;font-weight:800">
        🎯 Product Leadership Jobs
      </h1>
      <div style="color:#c7d2fe;font-size:12px;margin-top:6px">
        Personalised for Sujeet Singh · Health tech · BFSI · B2B SaaS · India &amp; Remote
      </div>
    </td>
  </tr>

  <!-- Stats bar -->
  <tr>
    <td style="background:#1e3a8a;padding:10px 28px;border-bottom:3px solid #3b82f6">
      <span style="color:#93c5fd;font-size:13px;font-weight:700">
        {count} roles found
      </span>
      &nbsp;<span style="color:#334155">·</span>&nbsp;
      <span style="font-size:12px">{source_summary}</span>
      &nbsp;&nbsp;
      <span style="color:#64748b;font-size:11px">sorted by relevancy ↓</span>
    </td>
  </tr>

  <!-- Excel attachment notice -->
  <tr>
    <td style="background:#fffbeb;padding:8px 28px;border-bottom:1px solid #fde68a;
        font-size:12px;color:#92400e;font-weight:600">
      📎 <strong>Full list attached as Excel</strong> — includes all roles, clickable URLs, score breakdown &amp; source summary.
      Email may show fewer rows due to display limits.
    </td>
  </tr>

  <!-- Legend -->
  <tr>
    <td style="background:#f0f9ff;padding:8px 28px;border-bottom:1px solid #bae6fd;
        font-size:11px;color:#475569">
      🏠 Remote &nbsp;·&nbsp; 🔀 Hybrid &nbsp;·&nbsp; 🏢 On-site
      &nbsp;&nbsp;|&nbsp;&nbsp;
      🇮🇳 India &nbsp;·&nbsp; 🇺🇸 USA &nbsp;·&nbsp; 🇬🇧 UK &nbsp;·&nbsp;
      🇸🇬 SG &nbsp;·&nbsp; 🌍 Global/Remote &nbsp;·&nbsp; 🌐 Unknown
      &nbsp;&nbsp;|&nbsp;&nbsp;
      <span style="background:#dcfce7;color:#166534;padding:1px 6px;
        border-radius:8px;font-weight:700">⭐ 8–10</span> Strong fit &nbsp;
      <span style="background:#fef9c3;color:#854d0e;padding:1px 6px;
        border-radius:8px;font-weight:700">⭐ 5–7</span> Possible &nbsp;
      <span style="background:#fee2e2;color:#991b1b;padding:1px 6px;
        border-radius:8px;font-weight:700">⭐ 1–4</span> Weak fit
    </td>
  </tr>

  <!-- Jobs table -->
  <tr>
    <td style="background:#fff;border-radius:0 0 12px 12px;overflow:hidden">
      <table width="100%" cellpadding="0" cellspacing="0">
        <thead>
          <tr style="background:#f8fafc;border-bottom:2px solid #e2e8f0">
            <th style="padding:10px 10px 10px 16px;color:#94a3b8;font-size:11px;
                font-weight:700;text-align:center;width:32px">#</th>
            <th style="padding:10px;color:#64748b;font-size:11px;font-weight:700;
                text-align:left;text-transform:uppercase;letter-spacing:0.5px">
                Role &amp; Company</th>
            <th style="padding:10px 8px;color:#64748b;font-size:11px;font-weight:700;
                text-align:center;text-transform:uppercase;letter-spacing:0.5px">
                Country</th>
            <th style="padding:10px 8px;color:#64748b;font-size:11px;font-weight:700;
                text-align:center;text-transform:uppercase;letter-spacing:0.5px">
                Mode</th>
            <th style="padding:10px 8px;color:#64748b;font-size:11px;font-weight:700;
                text-align:center;text-transform:uppercase;letter-spacing:0.5px">
                Fit</th>
            <th style="padding:10px 16px 10px 8px;color:#64748b;font-size:11px;
                font-weight:700;text-align:left;text-transform:uppercase;
                letter-spacing:0.5px">Source</th>
          </tr>
        </thead>
        <tbody>
          {rows if rows else no_jobs}
        </tbody>
      </table>
    </td>
  </tr>

  <!-- Footer -->
  <tr>
    <td style="padding:16px 0;text-align:center;color:#94a3b8;font-size:11px">
      Sent daily at 7 PM · 22 sources: LinkedIn · Remotive · Built In · Otta · Google Jobs
      · Naukri · iimjobs · Hirist · Indeed India · Glassdoor India
      · The Product Folks · Instahyre · Weekday · Cutshort
      · Michael Page · Stanton Chase · Heidrick &amp; Struggles · ABC Consultants
      · Propella · Longhouse · SutraHR · Korn Ferry India<br>
      <span style="color:#cbd5e1">sujeetkumar0809@gmail.com</span>
    </td>
  </tr>

</table>
</td></tr>
</table>
</body>
</html>"""


# ─────────────────────────────────────────────
# EMAIL SENDER
# ─────────────────────────────────────────────

# ─────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────

def build_excel(jobs: List[Dict]) -> Optional[bytes]:
    """Return xlsx bytes for all jobs, or None if openpyxl missing."""
    if not HAS_OPENPYXL:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Jobs {TODAY}"

    # ── Colour palette ────────────────────────────────────────────
    HDR_FILL   = PatternFill("solid", fgColor="1E3A5F")   # dark navy
    GREEN_FILL = PatternFill("solid", fgColor="DCFCE7")
    AMBER_FILL = PatternFill("solid", fgColor="FEF9C3")
    RED_FILL   = PatternFill("solid", fgColor="FEE2E2")
    ALT_FILL   = PatternFill("solid", fgColor="F8FAFC")   # light row stripe

    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    # ── Headers ───────────────────────────────────────────────────
    COLS = [
        ("#",          5),
        ("Title",      40),
        ("Company",    25),
        ("Location",   22),
        ("Country 🌍", 10),
        ("Mode 🏠",    10),
        ("Score ⭐",    9),
        ("Source",     16),
        ("URL",        55),
        ("Date",       12),
    ]
    for col_idx, (header, width) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"   # freeze header row
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    # ── Data rows ─────────────────────────────────────────────────
    for row_idx, j in enumerate(jobs, start=2):
        score     = j.get("_score", 5)
        wm_emoji  = j.get("_work_mode", ("🔀", "Hybrid"))[0]
        wm_label  = j.get("_work_mode", ("🔀", "Hybrid"))[1]
        flag      = j.get("_country_flag", "🌐")
        is_alt    = (row_idx % 2 == 0)

        row_data = [
            row_idx - 1,          # rank
            j.get("title", ""),
            j.get("company", ""),
            j.get("location", ""),
            flag,
            f"{wm_emoji} {wm_label}",
            score,
            j.get("source", ""),
            j.get("url", ""),
            j.get("date", ""),
        ]

        # Score cell fill
        if score >= 8:
            score_fill = GREEN_FILL
        elif score >= 5:
            score_fill = AMBER_FILL
        else:
            score_fill = RED_FILL

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == 2))
            # Row background
            if col_idx == 7:                         # Score column
                cell.fill      = score_fill
                cell.font      = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_idx in (5, 6):                  # flag / mode — centre
                cell.alignment = Alignment(horizontal="center", vertical="top")
                if is_alt: cell.fill = ALT_FILL
            elif col_idx == 1:                       # rank — centre
                cell.alignment = Alignment(horizontal="center", vertical="top")
                cell.font      = Font(color="94A3B8", bold=True)
                if is_alt: cell.fill = ALT_FILL
            elif col_idx == 9:                       # URL — make it a hyperlink
                if value and str(value).startswith("http"):
                    cell.hyperlink = str(value)
                    cell.value     = "🔗 Open"
                    cell.font      = Font(color="1D4ED8", underline="single")
                if is_alt: cell.fill = ALT_FILL
            else:
                if is_alt: cell.fill = ALT_FILL

        ws.row_dimensions[row_idx].height = 18

    # ── Summary tab ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Product Leadership Job Digest"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A2"] = f"Generated: {datetime.now().strftime('%A, %B %d %Y  %I:%M %p')}"
    ws2["A3"] = f"Total roles: {len(jobs)}"
    ws2["A4"] = f"Score 8–10 (strong fit): {sum(1 for j in jobs if j.get('_score',0)>=8)}"
    ws2["A5"] = f"Score 5–7 (good fit):    {sum(1 for j in jobs if 5<=j.get('_score',0)<8)}"
    ws2["A6"] = f"Score 1–4 (low fit):     {sum(1 for j in jobs if j.get('_score',0)<5)}"
    ws2.column_dimensions["A"].width = 45

    # Source breakdown
    ws2["A8"] = "Roles by source"
    ws2["A8"].font = Font(bold=True)
    by_src: Dict[str, int] = {}
    for j in jobs:
        by_src[j.get("source", "?")] = by_src.get(j.get("source", "?"), 0) + 1
    for r_off, (src, cnt) in enumerate(sorted(by_src.items(), key=lambda x: -x[1]), start=9):
        ws2.cell(row=r_off, column=1, value=src)
        ws2.cell(row=r_off, column=2, value=cnt)
        ws2.column_dimensions["B"].width = 8

    # Save to bytes buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def send_email(jobs: List[Dict], sources_used: List[str]) -> None:
    top_score = jobs[0]["_score"] if jobs else 0
    total     = len(jobs)
    date_str  = datetime.now().strftime("%b %d, %Y")

    subject = (
        f"🎯 {total} Product Jobs · Top {top_score}/10 — {date_str}"
    )

    # ── Outer wrapper: mixed (supports attachments) ───────────────
    msg            = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"]    = SENDER_EMAIL
    msg["To"]      = RECEIVER_EMAIL

    # ── Inner: alternative (plain + html versions) ────────────────
    alt = MIMEMultipart("alternative")

    plain = "\n".join(
        f"[{j['_score']}/10] {j['title']} @ {j['company']} "
        f"| {j.get('_country_flag','')} {j.get('location','')} "
        f"| {j['url']}"
        for j in jobs
    ) or "No roles found today."
    alt.attach(MIMEText(plain, "plain"))
    alt.attach(MIMEText(build_html(jobs, sources_used), "html"))
    msg.attach(alt)

    # ── Excel attachment (full list, all columns) ─────────────────
    xlsx_bytes = build_excel(jobs)
    if xlsx_bytes:
        fname = f"ProductJobs_{TODAY}.xlsx"
        part  = MIMEBase(
            "application",
            "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        part.set_payload(xlsx_bytes)
        email_encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{fname}"')
        part.add_header("Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            name=fname)
        msg.attach(part)
        print(f"  📎 Excel attachment ready: {fname} ({len(xlsx_bytes)//1024} KB)")

    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.ehlo()
        server.starttls()
        server.login(SENDER_EMAIL, GMAIL_APP_PASS)
        server.sendmail(SENDER_EMAIL, RECEIVER_EMAIL, msg.as_string())
    print(f"✅ Email sent — {total} jobs in Excel, top score {top_score}/10")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

SOURCES = [
    # ── Global ───────────────────────────────
    ("Remotive",             fetch_remotive),
    ("LinkedIn",             fetch_linkedin),
    ("Built In",             fetch_builtin),
    ("RemoteOK",             fetch_remoteok),      # replaced dead Otta
    ("TimesJobs",            fetch_timesjobs),     # replaced Google Jobs (blocked)
    # ── India mass-market ────────────────────
    ("Naukri",               fetch_naukri),
    ("iimjobs",              fetch_iimjobs),
    ("Shine.com",            fetch_shine),         # replaced JS-only Hirist
    ("Indeed India",         fetch_indeed_india),
    ("Foundit India",        fetch_foundit),       # replaced Glassdoor (requires JS+login)
    # ── India community / tech ───────────────
    ("The Product Folks",    fetch_tpf),
    ("Instahyre",            fetch_instahyre),
    ("Weekday",              fetch_weekday),
    ("Cutshort",             fetch_cutshort),
    # ── Executive search & boutique ──────────
    ("Michael Page India",   fetch_michael_page),
    ("Stanton Chase",        fetch_stanton_chase),
    ("Heidrick & Struggles", fetch_heidrick),
    ("ABC Consultants",      fetch_abc_consultants),
    ("Propella/Native",      fetch_propella),
    ("Longhouse",            fetch_longhouse),
    ("SutraHR",              fetch_sutrahr),
    ("Korn Ferry India",     fetch_korn_ferry_india),
]


def main() -> None:
    print(f"\n{'='*55}")
    print(f"  Product Leadership Job Digest — {TODAY}")
    print(f"{'='*55}\n")

    all_jobs: List[Dict]   = []
    sources_used: List[str] = []

    for name, fn in SOURCES:
        try:
            jobs = fn()
            if jobs:
                sources_used.append(name)
            all_jobs.extend(jobs)
        except Exception as e:
            print(f"  ✗ {name} crashed: {e}")

    all_jobs = deduplicate(all_jobs)

    # Enrich each job with score, flag, work-mode
    for j in all_jobs:
        j["_score"]        = score_job(j)
        j["_country_flag"] = detect_country_flag(j.get("location", ""))
        j["_work_mode"]    = detect_work_mode(j.get("title", ""), j.get("location", ""))

    # Sort: highest score first, then newest date
    all_jobs.sort(key=lambda j: (j["_score"], j.get("date", "")), reverse=True)

    print(f"\n{'─'*40}")
    print(f"Total unique roles : {len(all_jobs)}")
    if all_jobs:
        print(f"Score distribution : "
              f"8–10 → {sum(1 for j in all_jobs if j['_score']>=8)}  "
              f"5–7 → {sum(1 for j in all_jobs if 5<=j['_score']<8)}  "
              f"1–4 → {sum(1 for j in all_jobs if j['_score']<5)}")
    print(f"{'─'*40}\n")

    if not GMAIL_APP_PASS:
        print("⚠  GMAIL_APP_PASS not set in the environment — skipping send.\n")
        sys.exit(0)

    send_email(all_jobs, sources_used)


if __name__ == "__main__":
    main()
